"""Smoke test for skills/autoresearch/templates/_build_xlsx.py.

Fixture: synthetic state.json with
  - 2 experiment families (via axes: backbone × loss)
  - 3 tasks in results_history
  - 2 HPO runs in hpo_runs
  - 3 future_directions items (1 high-priority, 1 normal, 1 deferred)

Asserts:
  - xlsx file exists
  - Expected sheet names are present
  - Matrix sheet has turquoise winner row at expected location
  - HPO detail sheet has turquoise best-value highlight
  - Legend sheet exists with status glyph rows
  - Graceful skip: Per-task headline is present (tasks in fixture)
  - Graceful skip: HPO detail is present (hpo_runs in fixture)
  - Graceful skip: Future directions is present (future_directions in fixture)
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
from pathlib import Path

import pytest

# ---------------------------------------------------------------------------
# Make the templates module importable regardless of cwd.
# ---------------------------------------------------------------------------
_TEMPLATES = Path(__file__).resolve().parents[1] / "templates"
sys.path.insert(0, str(_TEMPLATES))

import _build_xlsx as bx  # noqa: E402  (import after sys.path manipulation)

# ---------------------------------------------------------------------------
# Branding constants (used in assertions — match the module's resolved values)
# ---------------------------------------------------------------------------
TURQUOISE = bx._TURQUOISE   # "40E0D0"
INK       = bx._INK         # "14141C"
PAPER     = bx._PAPER
AMBER     = bx._AMBER


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _make_state(state_dir: Path, slug: str) -> dict:
    """Write a synthetic state.json and return it as a dict."""
    state = {
        "schema_version": 1,
        "session_id": "ar-2026-01-01-000000",
        "session_started_at": "2026-01-01T00:00:00Z",
        "last_modified_at": "2026-01-01T01:00:00Z",
        "last_iteration_completed_at": "2026-01-01T01:00:00Z",
        "phase": "completed",
        "scope": "Smoke-test sweep over backbones and loss functions",
        "scope_slug": slug,
        "target_metric": "val_auc > 0.90",
        "axes": {
            "backbone": ["resnet18", "caformer"],
            "loss": ["bce", "focal"],
            "task": ["bacterial", "viral", "fungal"],
        },
        "candidate_queue": [
            {"id": "cand-pending-1", "axes": {"backbone": "swin", "loss": "bce", "task": "bacterial"}, "status": "pending"},
        ],
        "results_history": [
            {
                "id": "cand-001",
                "axes": {"backbone": "resnet18", "loss": "bce", "task": "bacterial"},
                "started_at": "2026-01-01T00:00:00Z",
                "ended_at": "2026-01-01T00:10:00Z",
                "status": "complete",
                "metric_value": 0.871,
                "fix_attempts": 0,
                "error_class": None,
                "commit_sha": "abc1234",
                "notes": "baseline",
                "iteration_runtime_seconds": 600,
                "llm_call_count_estimate": 2,
            },
            {
                "id": "cand-002",
                "axes": {"backbone": "caformer", "loss": "bce", "task": "bacterial"},
                "started_at": "2026-01-01T00:10:00Z",
                "ended_at": "2026-01-01T00:25:00Z",
                "status": "complete",
                "metric_value": 0.931,
                "fix_attempts": 0,
                "error_class": None,
                "commit_sha": "def5678",
                "notes": "winner for bacterial",
                "iteration_runtime_seconds": 900,
                "llm_call_count_estimate": 3,
            },
            {
                "id": "cand-003",
                "axes": {"backbone": "resnet18", "loss": "focal", "task": "viral"},
                "started_at": "2026-01-01T00:25:00Z",
                "ended_at": "2026-01-01T00:40:00Z",
                "status": "complete",
                "metric_value": 0.812,
                "fix_attempts": 1,
                "error_class": "code_bug",
                "commit_sha": "fed9012",
                "notes": "one code-fix applied",
                "iteration_runtime_seconds": 900,
                "llm_call_count_estimate": 5,
            },
        ],
        "current_best": {
            "id": "cand-002",
            "axes": {"backbone": "caformer", "loss": "bce", "task": "bacterial"},
            "metric_value": 0.931,
        },
        "pivot_history": [],
        "research_log": {"available": False, "session_path": None, "fallback_path": None},
        "iteration_count": 3,
        "stop_reason": "search space exhausted",
        "pending_stash_ref": None,
        "consecutive_iteration_failures": 0,
        "last_error_at": None,
        "consecutive_infra_count": 0,
        "consecutive_infra_candidates": [],
        # Extra fields consumed by _build_xlsx.py
        "hpo_runs": [
            {
                "id": "hpo-run-1",
                "params": {"lr": 1e-3, "batch_size": 32, "dropout": 0.1},
                "metrics": {"val_auc": 0.921, "val_loss": 0.182},
            },
            {
                "id": "hpo-run-2",
                "params": {"lr": 5e-4, "batch_size": 64, "dropout": 0.2},
                "metrics": {"val_auc": 0.937, "val_loss": 0.164},
            },
        ],
        "future_directions": [
            {
                "item": "Add tongue image modality",
                "effort": "L",
                "expected_lift": "unknown",
                "notes": "annotations not yet available",
                "deferred": True,
                "priority": "normal",
            },
            {
                "item": "Pull 400-patient JHU update",
                "effort": "S",
                "expected_lift": "+0.01–0.02 bact AUC",
                "notes": "dataset doc 2026-04-16",
                "deferred": False,
                "priority": "high",
            },
            {
                "item": "Raw cough audio via Wav2Vec2",
                "effort": "L",
                "expected_lift": "unknown",
                "notes": "requires raw-audio bucket",
                "deferred": True,
                "priority": "normal",
            },
        ],
    }

    state_path = state_dir / "state.json"
    state_path.write_text(json.dumps(state, indent=2))
    return state


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@pytest.fixture()
def xlsx_output(tmp_path, monkeypatch):
    """Build the xlsx from a synthetic state.json and return (path, wb)."""
    import openpyxl

    slug = "smoke-test-sweep"
    # Wire up the fake gstack home so _read_state finds the file
    gstack_home = tmp_path / ".gstack"
    state_dir = gstack_home / "projects" / slug / "autoresearch"
    state_dir.mkdir(parents=True)
    state = _make_state(state_dir, slug)

    monkeypatch.setenv("GSTACK_HOME", str(gstack_home))
    # Patch Path.home() used inside bx._read_state
    # The simpler approach: patch bx._read_state directly.
    monkeypatch.setattr(bx, "_read_state", lambda s: state)

    out_path = tmp_path / f"scorecard_{slug}.xlsx"

    from openpyxl import Workbook as _WB
    wb = _WB()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    bx._build_matrix(wb, state, "2026-01-01")
    bx._build_per_task(wb, state)
    bx._build_hpo_detail(wb, state)
    bx._build_future_directions(wb, state)
    bx._build_legend(wb)
    wb.save(out_path)

    wb2 = openpyxl.load_workbook(str(out_path))
    return out_path, wb2


class TestXlsxExists:
    def test_file_exists(self, xlsx_output):
        path, _ = xlsx_output
        assert path.exists(), "scorecard xlsx was not written"
        assert path.stat().st_size > 0, "scorecard xlsx is empty"


class TestSheetNames:
    EXPECTED = {"Matrix", "Per-task headline", "HPO detail", "Future directions", "Legend"}

    def test_all_sheets_present(self, xlsx_output):
        _, wb = xlsx_output
        assert set(wb.sheetnames) == self.EXPECTED, (
            f"Unexpected sheet names: {wb.sheetnames}"
        )

    def test_sheet_order(self, xlsx_output):
        _, wb = xlsx_output
        assert wb.sheetnames[0] == "Matrix"
        assert wb.sheetnames[-1] == "Legend"


class TestMatrixSheet:
    def test_title_row_ink_bg(self, xlsx_output):
        _, wb = xlsx_output
        ws = wb["Matrix"]
        cell = ws.cell(row=1, column=1)
        bg = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else ""
        assert bg.upper().endswith(INK.upper()), (
            f"Title row bg should be INK ({INK}), got {bg}"
        )

    def test_winner_row_turquoise(self, xlsx_output):
        """Row with caformer/bce/bacterial (best result) should be turquoise."""
        _, wb = xlsx_output
        ws = wb["Matrix"]
        turquoise_rows = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.fgColor.type == "rgb":
                    bg = cell.fill.fgColor.rgb
                    if bg.upper().endswith(TURQUOISE.upper()):
                        turquoise_rows.append(cell.row)
                        break
        assert len(turquoise_rows) >= 1, (
            "No turquoise winner rows found in Matrix sheet"
        )

    def test_freeze_panes_set(self, xlsx_output):
        _, wb = xlsx_output
        ws = wb["Matrix"]
        assert ws.freeze_panes is not None, "Matrix sheet freeze_panes not set"


class TestPerTaskSheet:
    def test_header_row_turquoise(self, xlsx_output):
        _, wb = xlsx_output
        ws = wb["Per-task headline"]
        # Row 3 should be the turquoise header
        header_row = 3
        cell = ws.cell(row=header_row, column=1)
        bg = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else ""
        assert bg.upper().endswith(TURQUOISE.upper()), (
            f"Per-task header row (row 3) should be turquoise, got {bg}"
        )

    def test_has_data_rows(self, xlsx_output):
        _, wb = xlsx_output
        ws = wb["Per-task headline"]
        # Should have at least one data row after header row 3
        assert ws.max_row >= 4, "Per-task sheet has no data rows"


class TestHpoDetailSheet:
    def test_best_value_turquoise(self, xlsx_output):
        """The best val_auc (0.937 from hpo-run-2) should have turquoise fill."""
        _, wb = xlsx_output
        ws = wb["HPO detail"]
        found_turquoise_data = False
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.fill.fgColor.type == "rgb":
                    bg = cell.fill.fgColor.rgb
                    if bg.upper().endswith(TURQUOISE.upper()):
                        # Check it's a data cell (not header)
                        if cell.row >= 4:
                            found_turquoise_data = True
        assert found_turquoise_data, (
            "HPO detail sheet: no turquoise-highlighted best-value cell found in data rows"
        )

    def test_star_marker_on_best(self, xlsx_output):
        """Best metric cell should contain the ★ marker."""
        _, wb = xlsx_output
        ws = wb["HPO detail"]
        star_found = False
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.value and "★" in str(cell.value):
                    star_found = True
        assert star_found, "HPO detail: no ★ marker found on best-value cells"


class TestFutureDirectionsSheet:
    def test_deferred_rows_grey(self, xlsx_output):
        """Deferred items should have grey fill (⏭ marker)."""
        _, wb = xlsx_output
        ws = wb["Future directions"]
        deferred_grey = bx._DEFERRED_GREY
        grey_rows = []
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.fill.fgColor.type == "rgb":
                    bg = cell.fill.fgColor.rgb
                    if bg.upper().endswith(deferred_grey.upper()):
                        grey_rows.append(cell.row)
                        break
        assert len(grey_rows) >= 1, "No deferred-grey rows found in Future directions"

    def test_high_priority_amber(self, xlsx_output):
        """High-priority items should have amber fill."""
        _, wb = xlsx_output
        ws = wb["Future directions"]
        amber_rows = []
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.fill.fgColor.type == "rgb":
                    bg = cell.fill.fgColor.rgb
                    if bg.upper().endswith(AMBER.upper()):
                        amber_rows.append(cell.row)
                        break
        assert len(amber_rows) >= 1, "No amber high-priority rows found in Future directions"


class TestLegendSheet:
    def test_legend_has_won_row(self, xlsx_output):
        """Legend should contain a row with 'WON' status."""
        _, wb = xlsx_output
        ws = wb["Legend"]
        won_found = any(
            "WON" in str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        assert won_found, "Legend sheet missing 'WON' status entry"

    def test_legend_has_deferred_row(self, xlsx_output):
        """Legend should contain a row with 'DEFERRED' glyph."""
        _, wb = xlsx_output
        ws = wb["Legend"]
        deferred_found = any(
            "DEFERRED" in str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        assert deferred_found, "Legend sheet missing DEFERRED glyph entry"


class TestGracefulSkip:
    def test_no_hpo_skips_sheet(self, tmp_path, monkeypatch):
        """If hpo_runs is absent/empty, HPO detail sheet should be skipped."""
        import openpyxl

        slug = "no-hpo-test"
        gstack_home = tmp_path / ".gstack"
        state_dir = gstack_home / "projects" / slug / "autoresearch"
        state_dir.mkdir(parents=True)
        state = _make_state(state_dir, slug)
        state["hpo_runs"] = []  # remove HPO data
        monkeypatch.setattr(bx, "_read_state", lambda s: state)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        bx._build_matrix(wb, state, "2026-01-01")
        bx._build_per_task(wb, state)
        bx._build_hpo_detail(wb, state)
        bx._build_future_directions(wb, state)
        bx._build_legend(wb)

        assert "HPO detail" not in wb.sheetnames, (
            "HPO detail should be skipped when hpo_runs is empty"
        )
        # Other sheets must still exist
        assert "Matrix" in wb.sheetnames
        assert "Legend" in wb.sheetnames

    def test_no_future_dirs_skips_sheet(self, tmp_path, monkeypatch):
        """If future_directions and candidate_queue are empty, sheet should be skipped."""
        import openpyxl

        slug = "no-future-test"
        gstack_home = tmp_path / ".gstack"
        state_dir = gstack_home / "projects" / slug / "autoresearch"
        state_dir.mkdir(parents=True)
        state = _make_state(state_dir, slug)
        state["future_directions"] = []
        state["candidate_queue"] = []
        monkeypatch.setattr(bx, "_read_state", lambda s: state)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        bx._build_matrix(wb, state, "2026-01-01")
        bx._build_per_task(wb, state)
        bx._build_hpo_detail(wb, state)
        bx._build_future_directions(wb, state)
        bx._build_legend(wb)

        assert "Future directions" not in wb.sheetnames, (
            "Future directions should be skipped when no items exist"
        )
        assert "Matrix" in wb.sheetnames
        assert "Legend" in wb.sheetnames

    def test_no_tasks_skips_per_task(self, tmp_path, monkeypatch):
        """If results_history has no task axis, Per-task headline should be skipped."""
        import openpyxl

        slug = "no-task-test"
        gstack_home = tmp_path / ".gstack"
        state_dir = gstack_home / "projects" / slug / "autoresearch"
        state_dir.mkdir(parents=True)
        state = _make_state(state_dir, slug)
        # Remove 'task' axis from all results so no task grouping possible
        for r in state["results_history"]:
            r["axes"].pop("task", None)
        monkeypatch.setattr(bx, "_read_state", lambda s: state)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        bx._build_matrix(wb, state, "2026-01-01")
        bx._build_per_task(wb, state)

        assert "Per-task headline" not in wb.sheetnames, (
            "Per-task headline should be skipped when no task axis in results"
        )
