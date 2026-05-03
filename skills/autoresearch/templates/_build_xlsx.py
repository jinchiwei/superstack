"""Build a brand-styled experiment matrix scorecard xlsx for an autoresearch session.

Reads:
  ~/.gstack/projects/<slug>/autoresearch/state.json

Writes (default):
  results/<date>_<scope>/scorecard_<date>.xlsx

Generates up to 5 sheets (gracefully skipped when state data is absent):
  Sheet 1 — Matrix:            experiment × axis grid with winner rows
  Sheet 2 — Per-task headline: best config + key metrics per task/target
  Sheet 3 — HPO detail:        hyperparameter sweep results table
  Sheet 4 — Future directions: deferred / next-iteration items
  Sheet 5 — Legend:            color & status glyph reference

Brand palette from skills/_shared/branding.py:
  header bg:  INK   #141414, white text, bold
  sub-header: dark  #222222, light-grey text
  winner:     TURQUOISE #40E0D0, ink text, bold, 🏆 marker
  section:    TURQUOISE fill on section-header rows
  warning:    AMBER #F0C840, white text
  deferred:   light grey #E8E8E8, dark text, ⏭ marker
  body:       PAPER #FAFAFA fill, ink text
  tab color:  TURQUOISE on every sheet

Run:
    python skills/autoresearch/templates/_build_xlsx.py \\
        --date 2026-04-30 --scope fw-arch-sweep

    # explicit output path:
    python skills/autoresearch/templates/_build_xlsx.py \\
        --date 2026-04-30 --scope fw-arch-sweep \\
        --output /path/to/scorecard.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "openpyxl not installed. Run: pip install openpyxl"
    )

# ---------------------------------------------------------------------------
# Brand palette — pull from _shared/branding.py constants.
# We resolve the path relative to this file so the script works regardless of
# cwd.  Fall back to inline hex literals if the module isn't importable (e.g.
# during isolated unit testing with a minimal fixture).
# ---------------------------------------------------------------------------

_HERE = Path(__file__).resolve()
_SHARED = _HERE.parents[2] / "_shared"
sys.path.insert(0, str(_SHARED))

try:
    import branding  # type: ignore

    _TURQUOISE  = branding.TURQUOISE.lstrip("#")   # "40E0D0"
    _DEEPPINK   = branding.DEEPPINK.lstrip("#")    # "FF1493"
    _AMBER      = branding.AMBER.lstrip("#")       # "F0C840"
    _BLUEVIOLET = branding.BLUEVIOLET.lstrip("#")  # "8A2BE2"
    _INK        = branding.INK.lstrip("#")         # "14141C"
    _PAPER      = branding.PAPER.lstrip("#")       # "FAFAFC"
    _WHITE      = branding.WHITE.lstrip("#")       # "FFFFFF"
    _MUTED      = branding.MUTED.lstrip("#")       # "555560"
except ImportError:
    _TURQUOISE  = "40E0D0"
    _DEEPPINK   = "FF1493"
    _AMBER      = "F0C840"
    _BLUEVIOLET = "8A2BE2"
    _INK        = "14141C"
    _PAPER      = "FAFAFC"
    _WHITE      = "FFFFFF"
    _MUTED      = "555560"

# Extra neutrals used in the reference scorecard
_DARK_SUBHEADER = "222222"   # row-2 subtitle bg
_DEFERRED_GREY  = "E8E8E8"   # deferred row bg
_LIGHT_GREY_TXT = "CCCCCC"   # sub-header text on dark bg
_RAN_LIGHT      = "CDEFEB"   # "ran, not winner" light teal

# openpyxl fill/font type abbrevs
_SOLID = "solid"


# ---------------------------------------------------------------------------
# Fill / font helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type=_SOLID, fgColor=hex_color)


def _font(
    hex_color: str = _INK,
    *,
    bold: bool = False,
    italic: bool = False,
    size: int = 10,
    name: str = "Geist",
) -> Font:
    return Font(color=hex_color, bold=bold, italic=italic, size=size, name=name)


def _align(
    horizontal: str = "left",
    vertical: str = "center",
    wrap: bool = False,
) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _thin_border() -> Border:
    thin = Side(style="thin", color="DDDDDD")
    return Border(bottom=thin)


# ---------------------------------------------------------------------------
# Row-writing helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, row_num: int, values: list[str], *, col_offset: int = 1) -> None:
    """Dark ink header row (H1-style): ink bg, white bold text."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        c.fill = _fill(_INK)
        c.font = _font(_WHITE, bold=True)
        c.alignment = _align()


def _write_subheader_row(ws, row_num: int, value: str, n_cols: int) -> None:
    """Dark #222 sub-header row spanning n_cols, light-grey text."""
    c = ws.cell(row=row_num, column=1, value=value)
    c.fill = _fill(_DARK_SUBHEADER)
    c.font = _font(_LIGHT_GREY_TXT, size=9)
    c.alignment = _align()
    # Clear the rest of the row
    for col in range(2, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = _fill(_DARK_SUBHEADER)


def _write_section_row(ws, row_num: int, value: str, n_cols: int) -> None:
    """Turquoise section header row (like 'Backbone', 'Inputs')."""
    c = ws.cell(row=row_num, column=1, value=value)
    c.fill = _fill(_TURQUOISE)
    c.font = _font(_INK, bold=True)
    c.alignment = _align()
    for col in range(2, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = _fill(_TURQUOISE)


def _write_winner_row(ws, row_num: int, values: list[str], *, col_offset: int = 1) -> None:
    """Turquoise winner row: turquoise bg, ink text, bold, 🏆 prepended."""
    for i, val in enumerate(values):
        display = val
        if i == 0 and val and not val.startswith("🏆"):
            display = f"🏆 {val}"
        c = ws.cell(row=row_num, column=col_offset + i, value=display)
        c.fill = _fill(_TURQUOISE)
        c.font = _font(_INK, bold=True)
        c.alignment = _align()


def _write_deferred_row(ws, row_num: int, values: list[str], *, col_offset: int = 1) -> None:
    """Light-grey deferred row with ⏭ marker."""
    for i, val in enumerate(values):
        display = val
        if i == 0 and val and not val.startswith("⏭"):
            display = f"⏭ {val}"
        c = ws.cell(row=row_num, column=col_offset + i, value=display)
        c.fill = _fill(_DEFERRED_GREY)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)


def _write_warning_row(ws, row_num: int, values: list[str], *, col_offset: int = 1) -> None:
    """Amber warning row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        c.fill = _fill(_AMBER)
        c.font = _font(_INK, bold=True)
        c.alignment = _align(wrap=True)


def _write_body_row(ws, row_num: int, values: list[str], *, col_offset: int = 1) -> None:
    """Standard paper-colored body row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)


def _write_metric_cell(ws, row_num: int, col_num: int, value: str) -> None:
    """Blueviolet metric value cell."""
    c = ws.cell(row=row_num, column=col_num, value=value)
    c.fill = _fill(_PAPER)
    c.font = _font(_BLUEVIOLET, bold=True, name="Geist Mono")
    c.alignment = _align(horizontal="center")


def _set_tab_color(ws) -> None:
    ws.sheet_properties.tabColor = _TURQUOISE


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# State reading
# ---------------------------------------------------------------------------

def _read_state(slug: str) -> dict:
    gstack_home = Path.home() / ".gstack"
    state_path = gstack_home / "projects" / slug / "autoresearch" / "state.json"
    if not state_path.exists():
        raise SystemExit(f"No state.json found at {state_path}")
    with state_path.open() as fh:
        return json.load(fh)


def _humanize(slug: str) -> str:
    if not slug:
        return ""
    slug = re.sub(r"^iter-\d+_", "", slug)
    parts = re.split(r"[-_]+", slug)
    return " ".join(p[:1].upper() + p[1:] if p else "" for p in parts).strip()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_matrix(wb: Workbook, state: dict, date_str: str) -> None:
    """Sheet 1 — Matrix: experiment grid with axes × experiment families."""
    scope_slug = state.get("scope_slug", "")
    axes: dict = state.get("axes", {})
    results_history: list[dict] = state.get("results_history", [])
    current_best: dict | None = state.get("current_best")

    ws = wb.create_sheet("Matrix")
    _set_tab_color(ws)

    # Experiment families: derived from axes keys (or from candidate IDs)
    # We use axes keys as the "dimension" rows and the unique axis values as
    # per-family columns.  For simplicity each axis key = one row group.
    axis_names = list(axes.keys()) if axes else []

    # Collect experiment family names from results_history candidate IDs
    exp_families: list[str] = []
    seen: set[str] = set()
    for r in results_history:
        r_axes: dict = r.get("axes", {})
        # Family = first axis value combo as a short label
        fam = "_".join(str(v) for v in list(r_axes.values())[:2]) if r_axes else r.get("id", "")
        if fam and fam not in seen:
            exp_families.append(fam)
            seen.add(fam)

    n_families = max(len(exp_families), 1)
    n_cols = 2 + n_families   # col A = #/idx, col B = Item, col C..N = families

    # Row 1 — title
    title_text = f"{_humanize(scope_slug)} Experiment Matrix — {date_str}"
    c = ws.cell(row=1, column=1, value=title_text)
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    c.alignment = _align()
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    # Row 2 — sub-header: scope text
    sub = state.get("scope", "") or _humanize(scope_slug)
    _write_subheader_row(ws, 2, sub, n_cols)

    # Row 3 — blank spacer
    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = _fill("FFFFFF")

    # Row 4 — column headers: #, Item, then one col per experiment family
    headers = ["#", "Item"] + [_humanize(f) for f in exp_families]
    _write_header_row(ws, 4, headers)

    # Row 5 — winner summary row (deeppink bg, per-experiment winner)
    winner_row_vals = ["★", "Per-experiment winner"]
    best_axes: dict = (current_best.get("axes") or {}) if current_best else {}
    for fam in exp_families:
        # Find best result in this family
        fam_results = [r for r in results_history
                       if "_".join(str(v) for v in list(r.get("axes", {}).values())[:2]) == fam]
        best_in_fam = next(
            (r for r in sorted(fam_results,
                                key=lambda x: x.get("metric_value") or 0,
                                reverse=True)
             if r.get("status") == "complete"),
            None,
        )
        if best_in_fam:
            mv = best_in_fam.get("metric_value", "")
            winner_row_vals.append(f"🏆 WON  {mv}" if mv else "🏆 WON")
        else:
            winner_row_vals.append("—")

    for i, val in enumerate(winner_row_vals):
        c = ws.cell(row=5, column=i + 1, value=val)
        c.fill = _fill(_DEEPPINK)
        c.font = _font(_WHITE, bold=True)
        c.alignment = _align()

    # Rows 6..N — one section per axis
    current_row = 6
    idx = 1
    for ax_name, ax_values in axes.items():
        # Section header
        _write_section_row(ws, current_row, _humanize(ax_name), n_cols)
        current_row += 1

        for val in (ax_values if isinstance(ax_values, list) else [ax_values]):
            val_str = str(val)
            # Find results matching this axis/value combination
            matching = [
                r for r in results_history
                if r.get("axes", {}).get(ax_name) == val
            ]
            best_match = next(
                (r for r in sorted(matching,
                                   key=lambda x: x.get("metric_value") or 0,
                                   reverse=True)
                 if r.get("status") == "complete"),
                None,
            )
            is_winner = (
                best_axes.get(ax_name) == val
                if best_axes else False
            )
            # Per-family status cells
            fam_cells = []
            for fam in exp_families:
                fam_match = next(
                    (r for r in matching
                     if "_".join(str(v) for v in list(r.get("axes", {}).values())[:2]) == fam),
                    None,
                )
                if fam_match:
                    mv = fam_match.get("metric_value", "")
                    status = fam_match.get("status", "")
                    fam_cells.append(f"{status}  {mv}" if mv else status or "—")
                else:
                    fam_cells.append("")

            row_vals = [str(idx), val_str] + fam_cells
            if is_winner:
                _write_winner_row(ws, current_row, row_vals)
            else:
                _write_body_row(ws, current_row, row_vals)

            current_row += 1
            idx += 1

    # Column widths: A=4, B=30 (axis value), rest=20
    col_widths = [4, 30] + [20] * n_families
    _set_col_widths(ws, col_widths)

    # Freeze panes: below header at row 5 (after section header)
    ws.freeze_panes = "A5"


def _build_per_task(wb: Workbook, state: dict) -> None:
    """Sheet 2 — Per-task headline: best config + metrics per task/target."""
    results_history: list[dict] = state.get("results_history", [])
    target_metric: str = state.get("target_metric", "") or ""

    # Group results by task (requires explicit 'task' key in axes).
    # Falls back to a top-level 'task' field.  If neither is present the sheet
    # is skipped — we don't want spurious per-candidate rows when the project
    # doesn't have a multi-task structure.
    tasks: dict[str, list[dict]] = {}
    for r in results_history:
        task = r.get("axes", {}).get("task") or r.get("task") or ""
        if not task:
            continue
        tasks.setdefault(task, []).append(r)

    if not tasks:
        return  # skip sheet gracefully

    ws = wb.create_sheet("Per-task headline")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")
    sub = state.get("scope", "") or _humanize(scope_slug)

    # Row 1 title
    c = ws.cell(row=1, column=1, value=f"Per-task headline — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 8):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    # Row 2 sub-header
    _write_subheader_row(ws, 2, sub, 7)

    # Row 3 — column headers (turquoise)
    hdrs = ["Task", "Best config", target_metric or "Metric", "Status", "Iterations", "Best axes", "Notes"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.fill = _fill(_TURQUOISE)
        c.font = _font(_INK, bold=True)
        c.alignment = _align()

    # Data rows
    row_num = 4
    for task, task_results in tasks.items():
        # Best result = highest metric_value with status=complete
        completed = [r for r in task_results if r.get("status") == "complete"]
        if not completed:
            # use any result
            completed = task_results
        best = max(completed, key=lambda r: r.get("metric_value") or 0, default=None)
        if not best:
            continue

        mv = best.get("metric_value", "—")
        axes_str = ", ".join(f"{k}={v}" for k, v in (best.get("axes") or {}).items())
        candidate_id = best.get("id", "")
        fix_attempts = best.get("fix_attempts", 0)
        notes = best.get("notes", "") or ""

        row_vals = [
            _humanize(str(task)),
            _humanize(candidate_id),
            str(mv) if mv is not None else "—",
            best.get("status", "—"),
            str(len(task_results)),
            axes_str,
            notes,
        ]
        _write_body_row(ws, row_num, row_vals)
        # Override metric cell with blueviolet
        _write_metric_cell(ws, row_num, 3, str(mv) if mv is not None else "—")
        row_num += 1

    _set_col_widths(ws, [18, 22, 12, 12, 10, 30, 30])
    _freeze(ws)


def _build_hpo_detail(wb: Workbook, state: dict) -> None:
    """Sheet 3 — HPO detail: hyperparameter sweep results.

    Expects state.json to carry a 'hpo_runs' key:
      [{"id": "...", "params": {...}, "metrics": {...}}, ...]
    Gracefully skipped if absent or empty.
    """
    hpo_runs: list[dict] = state.get("hpo_runs") or []
    if not hpo_runs:
        return

    ws = wb.create_sheet("HPO detail")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")
    target_metric: str = state.get("target_metric", "") or ""

    c = ws.cell(row=1, column=1, value=f"HPO detail — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)

    # Collect all param keys and metric keys from runs
    param_keys: list[str] = []
    metric_keys: list[str] = []
    seen_p: set[str] = set()
    seen_m: set[str] = set()
    for run in hpo_runs:
        for k in (run.get("params") or {}).keys():
            if k not in seen_p:
                param_keys.append(k)
                seen_p.add(k)
        for k in (run.get("metrics") or {}).keys():
            if k not in seen_m:
                metric_keys.append(k)
                seen_m.add(k)

    all_cols = ["Run"] + param_keys + metric_keys
    n_cols = len(all_cols)

    # Title row tail
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    _write_subheader_row(ws, 2, state.get("scope", "") or "", n_cols)

    # Header row (turquoise)
    for i, h in enumerate(all_cols):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.fill = _fill(_TURQUOISE)
        c.font = _font(_INK, bold=True)
        c.alignment = _align(horizontal="center")

    # Find best value per metric column for conditional highlight
    best_metric_vals: dict[str, float] = {}
    for mk in metric_keys:
        vals = [run.get("metrics", {}).get(mk) for run in hpo_runs
                if run.get("metrics", {}).get(mk) is not None]
        numeric = [v for v in vals if isinstance(v, (int, float))]
        if numeric:
            best_metric_vals[mk] = max(numeric)

    row_num = 4
    for run in hpo_runs:
        run_id = run.get("id", f"run-{row_num - 3}")
        params = run.get("params") or {}
        metrics = run.get("metrics") or {}

        # Build row values
        row_vals = [run_id] + [str(params.get(k, "—")) for k in param_keys]
        c_start = len(row_vals) + 1  # 1-indexed column where metrics start

        _write_body_row(ws, row_num, row_vals)

        # Write metric cells — highlight best with turquoise
        for mi, mk in enumerate(metric_keys):
            mv = metrics.get(mk)
            col_num = c_start + mi
            if mv is None:
                ws.cell(row=row_num, column=col_num, value="—")
            else:
                is_best = (isinstance(mv, (int, float))
                           and best_metric_vals.get(mk) == mv)
                c = ws.cell(row=row_num, column=col_num, value=mv)
                if is_best:
                    c.fill = _fill(_TURQUOISE)
                    c.font = _font(_INK, bold=True)
                    # Add ★ marker as suffix in display
                    c.value = f"{mv} ★"
                else:
                    c.fill = _fill(_PAPER)
                    c.font = _font(_BLUEVIOLET, bold=False, name="Geist Mono")
                c.alignment = _align(horizontal="center")

        row_num += 1

    param_widths = [14] + [14] * len(param_keys) + [14] * len(metric_keys)
    _set_col_widths(ws, param_widths)
    _freeze(ws)


def _build_future_directions(wb: Workbook, state: dict) -> None:
    """Sheet 4 — Future directions: deferred / next-iteration items.

    Expects state.json to carry a 'future_directions' key:
      [{"item": "...", "effort": "S/M/L", "expected_lift": "...",
        "notes": "...", "priority": "high/normal", "deferred": true/false}, ...]
    Or falls back to candidate_queue items with status=pending that look like
    deferred work.

    Gracefully skipped if no data.
    """
    future: list[dict] = state.get("future_directions") or []

    # Fallback: pending queue items
    if not future:
        for c in state.get("candidate_queue") or []:
            if c.get("status") == "pending":
                axes_str = ", ".join(
                    f"{k}={v}" for k, v in (c.get("axes") or {}).items()
                )
                future.append({
                    "item": axes_str or c.get("id", ""),
                    "effort": "?",
                    "expected_lift": "",
                    "notes": "",
                    "deferred": True,
                })

    if not future:
        return

    ws = wb.create_sheet("Future directions")
    _set_tab_color(ws)

    scope_slug = state.get("scope_slug", "")

    c = ws.cell(row=1, column=1, value=f"Future directions — {_humanize(scope_slug)}")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 6):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    _write_subheader_row(ws, 2, state.get("scope", "") or "", 5)

    # Header row
    _write_header_row(ws, 3, ["#", "Experiment", "Effort", "Expected lift / value", "Notes / unlock"])

    row_num = 4
    for i, item in enumerate(future, start=1):
        label = item.get("item", "")
        effort = item.get("effort", "?")
        lift = item.get("expected_lift", "")
        notes = item.get("notes", "")
        deferred = item.get("deferred", True)
        priority = item.get("priority", "normal")

        row_vals = [str(i), label, effort, lift, notes]

        if priority == "high":
            _write_warning_row(ws, row_num, row_vals)
        elif deferred:
            _write_deferred_row(ws, row_num, row_vals)
        else:
            _write_body_row(ws, row_num, row_vals)

        row_num += 1

    _set_col_widths(ws, [4, 40, 8, 30, 30])
    _freeze(ws)


def _build_legend(wb: Workbook) -> None:
    """Sheet 5 — Legend: color and status glyph reference."""
    ws = wb.create_sheet("Legend")
    _set_tab_color(ws)

    # Row 1 title
    c = ws.cell(row=1, column=1, value="Legend — color & status reference")
    c.fill = _fill(_INK)
    c.font = _font(_WHITE, bold=True, size=12)
    for col in range(2, 5):
        ws.cell(row=1, column=col).fill = _fill(_INK)

    # Blank row 2
    ws.cell(row=2, column=1).value = ""

    # Section: Status codes
    row = 3
    c = ws.cell(row=row, column=1, value="Status codes")
    c.fill = _fill(_DEEPPINK)
    c.font = _font(_WHITE, bold=True)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_DEEPPINK)

    row += 1
    _write_header_row(ws, row, ["", "Status", "Meaning", "Sample fill"])

    statuses = [
        ("WON",      "Ran this iteration and was the winner for its experiment",     "🏆 WON",      _TURQUOISE),
        ("RAN",      "Ran this iteration; informative but not the top choice",       "✅ RAN",       _RAN_LIGHT),
        ("LOST",     "Ran and was ACTIVELY HARMFUL — avoid in future",               "❌ LOST",      _DEEPPINK),
        ("NEW",      "Added mid-session; added on top of the original plan",         "🆕 NEW",       _AMBER),
        ("DEF",      "Matrix item was never attempted (deferred to next sprint)",    "⏭ DEFERRED",  _DEFERRED_GREY),
        ("HALTED",   "Session halted due to infra failure",                          "🛑 HALTED",    _MUTED),
    ]

    for stat, meaning, sample, bg_hex in statuses:
        row += 1
        ws.cell(row=row, column=1).fill = _fill(_PAPER)
        c = ws.cell(row=row, column=2, value=stat)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c = ws.cell(row=row, column=3, value=meaning)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)
        c = ws.cell(row=row, column=4, value=sample)
        c.fill = _fill(bg_hex)
        c.font = _font(_INK, bold=(bg_hex in (_TURQUOISE, _AMBER, _DEEPPINK)))
        c.alignment = _align(horizontal="center")

    # Blank
    row += 2

    # Section: Color palette
    c = ws.cell(row=row, column=1, value="Color palette — priority order")
    c.fill = _fill(_DEEPPINK)
    c.font = _font(_WHITE, bold=True)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_DEEPPINK)

    row += 1
    palette = [
        ("#1", "Turquoise",  "headline metric · winner callouts · section headers",    _TURQUOISE),
        ("#2", "Deeppink",   "secondary accents · comparison points · LOST status",    _DEEPPINK),
        ("#3", "Amber",      "tertiary — NEW status · warning / high-priority rows",   _AMBER),
        ("#4", "Blueviolet", "metric values · quaternary callouts",                    _BLUEVIOLET),
        ("#5", "Paper",      "standard body cell fill",                                _PAPER),
        ("#6", "Ink",        "header / title backgrounds",                             _INK),
        ("#7", "Deferred",   "deferred / pending items",                               _DEFERRED_GREY),
    ]
    for rank, name, desc, bg_hex in palette:
        c = ws.cell(row=row, column=1, value=rank)
        c.fill = _fill(_PAPER)
        c.font = _font(_MUTED)
        c = ws.cell(row=row, column=2, value=name)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c = ws.cell(row=row, column=3, value=desc)
        c.fill = _fill(_PAPER)
        c.font = _font(_INK)
        c.alignment = _align(wrap=True)
        c = ws.cell(row=row, column=4, value=f"#{bg_hex}")
        c.fill = _fill(bg_hex)
        c.font = _font(_WHITE if bg_hex in (_INK, _BLUEVIOLET, _DEEPPINK) else _INK)
        c.alignment = _align(horizontal="center")
        row += 1

    # Final note row
    row += 1
    c = ws.cell(row=row, column=1,
                value="Rebuild: python skills/autoresearch/templates/_build_xlsx.py --scope <slug>")
    c.fill = _fill(_INK)
    c.font = _font(_MUTED, size=9)
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = _fill(_INK)

    _set_col_widths(ws, [6, 14, 48, 20])
    _freeze(ws, "A2")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--date",
        default=datetime.now().strftime("%Y-%m-%d"),
        help="Session date (YYYY-MM-DD). Default: today.",
    )
    parser.add_argument("--scope", required=True, help="Session scope slug.")
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output path for the xlsx. "
            "Default: results/<date>_<scope>/scorecard_<date>.xlsx"
        ),
    )
    args = parser.parse_args()

    state = _read_state(args.scope)

    if args.output:
        out = Path(args.output)
    else:
        out_dir = Path("results") / f"{args.date}_{args.scope}"
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / f"scorecard_{args.date}.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default "Sheet" created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_matrix(wb, state, args.date)
    _build_per_task(wb, state)
    _build_hpo_detail(wb, state)
    _build_future_directions(wb, state)
    _build_legend(wb)

    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
