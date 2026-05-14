"""Shared xlsx styling primitives for Jin's brand palette.

All build-xlsx and autoresearch xlsx rendering import from here.
Provides openpyxl helpers for fills, fonts, borders, alignment, and cell
styling presets (winner, deferred, warning, headline, header, body, title bar).
Also exposes add_glyph_to_cell() which embeds an FA icon PNG anchored to a
cell via openpyxl.drawing.image.Image.

Color constants are imported from skills/_shared/branding.py.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

# ---------------------------------------------------------------------------
# Import branding palette.  Resolved relative to this file so the module
# works regardless of cwd.
# ---------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

try:
    import branding  # type: ignore
    TURQUOISE   = branding.TURQUOISE.lstrip("#")   # "40E0D0"
    DEEPPINK    = branding.DEEPPINK.lstrip("#")    # "FF1493"
    AMBER       = branding.AMBER.lstrip("#")       # "F0C840"
    BLUEVIOLET  = branding.BLUEVIOLET.lstrip("#")  # "8A2BE2"
    INK         = branding.INK.lstrip("#")         # "14141C"
    PAPER       = branding.PAPER.lstrip("#")       # "FAFAFC"
    WHITE       = branding.WHITE.lstrip("#")       # "FFFFFF"
    MUTED       = branding.MUTED.lstrip("#")       # "555560"
    DIM         = branding.DIM.lstrip("#")         # "888888"
except ImportError:
    TURQUOISE   = "40E0D0"
    DEEPPINK    = "FF1493"
    AMBER       = "F0C840"
    BLUEVIOLET  = "8A2BE2"
    INK         = "14141C"
    PAPER       = "FAFAFC"
    WHITE       = "FFFFFF"
    MUTED       = "555560"
    DIM         = "888888"

# Extra neutrals not in branding.py
DEFERRED_GREY  = "E8E8E8"   # deferred row background
DARK_SUBHEADER = "222222"   # sub-header row background
LIGHT_GREY_TXT = "CCCCCC"   # text on dark sub-header rows
RAN_LIGHT      = "CDEFEB"   # "ran, not winner" light teal

# Solid fill type shorthand
_SOLID = "solid"

# ---------------------------------------------------------------------------
# Lazy openpyxl imports — we do not import at module level so that importing
# this module in environments without openpyxl just to use the color constants
# does not immediately fail.
# ---------------------------------------------------------------------------
try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> "PatternFill":
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type=_SOLID, fgColor=hex_color)


def _font(
    hex_color: str = INK,
    *,
    bold: bool = False,
    italic: bool = False,
    size: int = 11,
    name: str = "Geist",
) -> "Font":
    from openpyxl.styles import Font
    return Font(color=hex_color, bold=bold, italic=italic, size=size, name=name)


def _align(
    horizontal: str = "left",
    vertical: str = "center",
    wrap: bool = False,
) -> "Alignment":
    from openpyxl.styles import Alignment
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _thin_border() -> "Border":
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="DDDDDD")
    return Border(bottom=thin)


# ---------------------------------------------------------------------------
# Cell-level style appliers
# ---------------------------------------------------------------------------

def _apply_header_style(cell, *, col_count: int = 0) -> None:
    """Ink bg, white bold 11pt Geist Mono text. Standard column header."""
    cell.fill = _fill(INK)
    cell.font = _font(WHITE, bold=True, size=11, name="Geist Mono")
    cell.alignment = _align()


def _apply_body_style(cell, *, alt: bool = False) -> None:
    """Paper (or white for alternating) bg, ink text, 11pt Geist."""
    bg = PAPER if not alt else WHITE
    cell.fill = _fill(bg)
    cell.font = _font(INK, size=11)
    cell.alignment = _align(wrap=True)


def _apply_winner_style(cell) -> None:
    """Turquoise fill, ink bold text. For winner / best rows."""
    cell.fill = _fill(TURQUOISE)
    cell.font = _font(INK, bold=True, size=11)
    cell.alignment = _align()


def _apply_deferred_style(cell) -> None:
    """Light grey fill, dim ink text. For deferred / skipped items."""
    cell.fill = _fill(DEFERRED_GREY)
    cell.font = _font(INK, size=11)
    cell.alignment = _align(wrap=True)


def _apply_warning_style(cell) -> None:
    """Amber fill, ink bold text. For warnings / high-priority items."""
    cell.fill = _fill(AMBER)
    cell.font = _font(INK, bold=True, size=11)
    cell.alignment = _align(wrap=True)


def _apply_headline_style(cell) -> None:
    """Deeppink fill, white bold text. For headline / accent rows."""
    cell.fill = _fill(DEEPPINK)
    cell.font = _font(WHITE, bold=True, size=11)
    cell.alignment = _align()


def _apply_title_bar_style(cell) -> None:
    """Ink fill, white bold 12pt text. For the title bar at row 1."""
    cell.fill = _fill(INK)
    cell.font = _font(WHITE, bold=True, size=12)
    cell.alignment = _align()


def _apply_callout_style(cell, *, accent_hex: str = TURQUOISE) -> None:
    """Dark ink fill, white text. For takeaway callout below a table.

    Accent color is used for the bold prefix (e.g. "Takeaway:") if the
    caller wants to highlight it, but at the cell level we just style
    the full row dark; the caller can apply rich text separately if needed.
    """
    cell.fill = _fill(INK)
    cell.font = _font(WHITE, size=11)
    cell.alignment = _align(horizontal="left", vertical="center", wrap=True)


# ---------------------------------------------------------------------------
# Row-level helpers (write values + style in one call)
# ---------------------------------------------------------------------------

def _write_header_row(ws, row_num: int, values: list, *, col_offset: int = 1) -> None:
    """Write a full ink-bg/white-bold header row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_header_style(c)


def _write_body_row(
    ws,
    row_num: int,
    values: list,
    *,
    col_offset: int = 1,
    alt: bool = False,
) -> None:
    """Write a standard body row (paper fill, ink text)."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_body_style(c, alt=alt)


def _write_winner_row(ws, row_num: int, values: list, *, col_offset: int = 1) -> None:
    """Write a turquoise winner row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_winner_style(c)


def _write_deferred_row(ws, row_num: int, values: list, *, col_offset: int = 1) -> None:
    """Write a light-grey deferred row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_deferred_style(c)


def _write_warning_row(ws, row_num: int, values: list, *, col_offset: int = 1) -> None:
    """Write an amber warning row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_warning_style(c)


def _write_headline_row(ws, row_num: int, values: list, *, col_offset: int = 1) -> None:
    """Write a deeppink headline row."""
    for i, val in enumerate(values):
        c = ws.cell(row=row_num, column=col_offset + i, value=val)
        _apply_headline_style(c)


# ---------------------------------------------------------------------------
# Sheet-level helpers
# ---------------------------------------------------------------------------

def _set_tab_color(ws, color_hex: str = TURQUOISE) -> None:
    """Set the sheet tab color."""
    ws.sheet_properties.tabColor = color_hex


# Named tab colors recognised by `<!-- tab: <name> -->` markdown directive.
# Maps name → hex (without leading #). Add more here as needed.
TAB_COLORS = {
    "turquoise":   TURQUOISE,
    "deeppink":    DEEPPINK,
    "pink":        DEEPPINK,
    "amber":       AMBER,
    "yellow":      AMBER,
    "blueviolet":  BLUEVIOLET,
    "violet":      BLUEVIOLET,
    "purple":      BLUEVIOLET,
    "ink":         INK,
    "dark":        INK,
    "grey":        DEFERRED_GREY,
    "gray":        DEFERRED_GREY,
}


def resolve_tab_color(name_or_hex: str) -> str:
    """Resolve a tab-color spec to a 6-char hex string.

    Accepts:
      - Named colors: "turquoise", "deeppink", "amber", "blueviolet", "ink", ...
      - Hex with leading #: "#ff1493"
      - Hex without #: "ff1493"
    Returns a 6-char uppercase hex string (no #). Falls back to TURQUOISE
    if the input is unrecognised.
    """
    s = (name_or_hex or "").strip().lstrip("#").lower()
    if not s:
        return TURQUOISE
    if s in TAB_COLORS:
        return TAB_COLORS[s]
    # Accept raw hex (3 or 6 chars)
    if len(s) == 6 and all(c in "0123456789abcdef" for c in s):
        return s.upper()
    if len(s) == 3 and all(c in "0123456789abcdef" for c in s):
        return "".join(c * 2 for c in s).upper()
    return TURQUOISE


def _set_col_widths(ws, widths: list[float]) -> None:
    """Set column widths (list indexed from col A onward)."""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _auto_col_widths(ws, *, max_width: int = 40, min_width: int = 8) -> None:
    """Auto-size columns based on max content length, capped at max_width."""
    from openpyxl.utils import get_column_letter
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                length = len(str(cell.value))
                col_widths[cell.column] = max(col_widths.get(cell.column, 0), length)
    for col_idx, width in col_widths.items():
        clamped = max(min_width, min(width + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = clamped


def _freeze_header(ws, cell: str = "A2") -> None:
    """Freeze panes at `cell` (default: just below first data row)."""
    ws.freeze_panes = cell


def _build_title_bar(ws, title: str, n_cols: int, row_num: int = 1) -> None:
    """Write a full-width ink title bar at `row_num`.

    The first cell gets the title text; the rest of the row is filled with
    the same ink background so the bar spans all data columns.
    """
    c = ws.cell(row=row_num, column=1, value=title)
    _apply_title_bar_style(c)
    for col in range(2, n_cols + 1):
        ws.cell(row=row_num, column=col).fill = _fill(INK)


# ---------------------------------------------------------------------------
# FA glyph cell embedder
# ---------------------------------------------------------------------------

def add_glyph_to_cell(
    ws,
    cell,
    *,
    icon_name: str,
    color_hex: str,
    size_px: int = 24,
) -> bool:
    """Render an FA glyph in the brand color and anchor it to the cell.

    Reuses the shared icon registry (skills/_shared/icons/registry.py)
    to avoid duplicating SVG assets.  The rendered PNG is embedded as an
    openpyxl Image anchored at the cell's top-left corner.  Column width
    and row height are nudged to accommodate the glyph.

    Returns True if the glyph was successfully embedded, False if the icon
    could not be rendered (e.g. cairosvg not installed or icon not bundled).
    """
    try:
        from openpyxl.drawing.image import Image as XlImage
    except ImportError:
        return False

    # Resolve the shared icon registry.
    _icons_dir = Path(__file__).resolve().parent / "icons"
    _registry_path = _icons_dir / "registry.py"

    try:
        import importlib.util
        _spec = importlib.util.spec_from_file_location("_shared_icons_registry", str(_registry_path))
        _mod = importlib.util.module_from_spec(_spec)  # type: ignore[arg-type]
        _spec.loader.exec_module(_mod)  # type: ignore[union-attr]
        render_icon = _mod.render_icon
    except Exception:
        return False

    png_path = render_icon(icon_name, color_hex, size_px=size_px)
    if png_path is None:
        return False

    img = XlImage(str(png_path))
    # Set display size to size_px × size_px (EMU conversion: 1 px ≈ 9525 EMU at 96 dpi)
    emu = size_px * 9525
    img.width = size_px
    img.height = size_px

    # Anchor to the cell address (e.g. "A2")
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    img.anchor = cell.coordinate

    ws.add_image(img)

    # Adjust row height if needed (Excel row height in points; 1 px ≈ 0.75 pt)
    min_height_pt = size_px * 0.75 + 4  # 4pt padding
    row_dim = ws.row_dimensions[cell.row]
    if row_dim.height is None or row_dim.height < min_height_pt:
        row_dim.height = min_height_pt

    return True
