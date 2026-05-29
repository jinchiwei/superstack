"""build-xlsx: Markdown → brand-styled xlsx.

Converts a markdown document into a multi-sheet branded xlsx workbook.

Each H1 heading becomes one sheet.  Tables inside each section become
styled data rows.  Optional bracket-prefix markers apply semantic styling:
  [winner]   → turquoise fill + FaTrophy glyph
  [deferred] → light grey fill + FaForward glyph
  [warning]  → amber fill + FaTriangleExclamation glyph
  [headline] → deeppink fill + FaStar glyph

Usage:
  python build.py --input <md> --output <xlsx>
  python build.py --input <md> --output <xlsx> --no-title-bar --no-frozen-header --no-glyphs

See SKILL.md for full documentation.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Ensure shared modules are importable regardless of cwd.
# ---------------------------------------------------------------------------
_SKILL_DIR = Path(__file__).resolve().parent
_SKILLS_DIR = _SKILL_DIR.parent
_SHARED_DIR = _SKILLS_DIR / "_shared"

for _p in [str(_SHARED_DIR), str(_SKILLS_DIR)]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

# ---------------------------------------------------------------------------
# openpyxl
# ---------------------------------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Shared styling primitives
# ---------------------------------------------------------------------------
from branding_xlsx import (  # type: ignore
    TURQUOISE, DEEPPINK, AMBER, BLUEVIOLET, INK, PAPER, WHITE, MUTED,
    _fill, _font, _align,
    _write_header_row, _write_body_row,
    _write_winner_row, _write_deferred_row,
    _write_warning_row, _write_headline_row,
    _apply_callout_style,
    _set_tab_color, _set_col_widths, _auto_col_widths,
    _freeze_header, _build_title_bar,
    add_glyph_to_cell,
    resolve_tab_color,
)

# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------
from md_parser import (  # type: ignore
    parse_markdown,
    MARKER_WINNER, MARKER_DEFERRED, MARKER_WARNING, MARKER_HEADLINE,
    SheetSpec, TableSpec, ProseBlock, CalloutBlock,
)

# ---------------------------------------------------------------------------
# Marker → (style_fn, icon_name, icon_color)
# ---------------------------------------------------------------------------
_MARKER_CONFIG = {
    MARKER_WINNER:   (_write_winner_row,   "FaTrophy",              f"#{TURQUOISE}"),
    MARKER_DEFERRED: (_write_deferred_row, "FaForward",             f"#{MUTED}"),
    MARKER_WARNING:  (_write_warning_row,  "FaTriangleExclamation", f"#{AMBER}"),
    MARKER_HEADLINE: (_write_headline_row, "FaStar",                f"#{DEEPPINK}"),
}


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def _build_sheet(
    wb: Workbook,
    spec: SheetSpec,
    *,
    with_title_bar: bool,
    with_frozen_header: bool,
    with_glyphs: bool,
) -> None:
    """Render one SheetSpec into a new worksheet."""
    ws = wb.create_sheet(spec.name)
    if spec.tab_color:
        _set_tab_color(ws, resolve_tab_color(spec.tab_color))
    else:
        _set_tab_color(ws)

    current_row = 1
    last_table_n_cols = 1  # tracks col span for a CalloutBlock to span across

    # Title bar (row 1)
    if with_title_bar:
        # We need to know n_cols; we'll use a provisional value of 10,
        # then update after we know the actual column count.
        _title_row = current_row
        current_row += 1

    # Track max column count for retrospective title bar sizing
    max_cols_seen = 1

    for block in spec.blocks:
        if isinstance(block, TableSpec):
            n_cols = len(block.headers)
            max_cols_seen = max(max_cols_seen, n_cols)
            last_table_n_cols = n_cols

            # Header row
            _write_header_row(ws, current_row, block.headers)
            header_row = current_row
            current_row += 1

            # Freeze just below header (if first table on sheet)
            if with_frozen_header:
                _freeze_header(ws, f"A{current_row}")
                with_frozen_header = False  # only freeze once per sheet

            # Data rows
            for row_idx, row_cells in enumerate(block.rows):
                values = [cd.value for cd in row_cells]

                # Determine dominant marker for the row (use first cell's marker)
                row_marker = row_cells[0].marker if row_cells else None

                # Check if any cell in row has a marker (first cell takes precedence)
                # but also check all cells for per-cell markers
                alt = row_idx % 2 == 1

                if row_marker and row_marker in _MARKER_CONFIG:
                    style_fn, icon_name, icon_color = _MARKER_CONFIG[row_marker]
                    style_fn(ws, current_row, values)
                    if with_glyphs:
                        # Anchor the glyph in the first column PAST the table
                        # (white background) so the brand-colored icon is
                        # visible — rendering it over the same-colored fill in
                        # column 1 made it invisible and clipped the row text.
                        glyph_cell = ws.cell(row=current_row, column=n_cols + 1)
                        add_glyph_to_cell(
                            ws,
                            glyph_cell,
                            icon_name=icon_name,
                            color_hex=icon_color,
                            size_px=18,
                        )
                else:
                    _write_body_row(ws, current_row, values, alt=alt)

                    # Handle per-cell markers (non-first cells may also have markers)
                    for col_idx, cd in enumerate(row_cells):
                        if cd.marker and col_idx > 0 and cd.marker in _MARKER_CONFIG:
                            style_fn, _, _ = _MARKER_CONFIG[cd.marker]
                            cell = ws.cell(row=current_row, column=col_idx + 1)
                            # Re-apply cell-level style
                            if cd.marker == MARKER_WINNER:
                                from branding_xlsx import _apply_winner_style
                                _apply_winner_style(cell)
                            elif cd.marker == MARKER_DEFERRED:
                                from branding_xlsx import _apply_deferred_style
                                _apply_deferred_style(cell)
                            elif cd.marker == MARKER_WARNING:
                                from branding_xlsx import _apply_warning_style
                                _apply_warning_style(cell)
                            elif cd.marker == MARKER_HEADLINE:
                                from branding_xlsx import _apply_headline_style
                                _apply_headline_style(cell)

                current_row += 1

        elif isinstance(block, ProseBlock):
            # Prose: one cell per line in column A, paper bg, body font
            for line in block.text.splitlines():
                line = line.strip()
                if not line:
                    current_row += 1
                    continue
                c = ws.cell(row=current_row, column=1, value=line)
                from branding_xlsx import _apply_body_style
                _apply_body_style(c)
                current_row += 1

        elif isinstance(block, CalloutBlock):
            # Dark takeaway callout spanning the most-recent table's columns.
            n_cols = max(last_table_n_cols, 1)
            max_cols_seen = max(max_cols_seen, n_cols)
            # Render as a rich-text-style cell: "**LABEL:** body" in one cell.
            # openpyxl doesn't do mixed bold/regular within a single .value
            # without inline rich text on shared strings; for simplicity we
            # build a single string and rely on the dark callout fill + the
            # leading "LABEL:" prefix to signal emphasis.
            display = f"{block.label.upper()}: {block.text}" if block.label else block.text
            c = ws.cell(row=current_row, column=1, value=display)
            _apply_callout_style(c)
            if n_cols > 1:
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=n_cols,
                )
                # Apply fill to all spanned cells so the merge looks unified
                for col in range(2, n_cols + 1):
                    side_cell = ws.cell(row=current_row, column=col)
                    _apply_callout_style(side_cell)
            # Slightly taller row to fit wrapping callout text
            row_dim = ws.row_dimensions[current_row]
            if row_dim.height is None or row_dim.height < 32:
                row_dim.height = 32
            current_row += 1

    # Retrospectively fill the title bar now we know max_cols_seen
    if with_title_bar:
        _build_title_bar(ws, spec.name, max(max_cols_seen, 1), row_num=_title_row)

    # Auto-size columns
    _auto_col_widths(ws, max_width=40)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_xlsx(
    markdown_text: str,
    *,
    with_title_bar: bool = True,
    with_frozen_header: bool = True,
    with_glyphs: bool = True,
) -> Workbook:
    """Convert markdown text to a brand-styled Workbook.

    Returns an openpyxl Workbook (not yet saved).
    """
    result = parse_markdown(markdown_text)

    wb = Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for spec in result.sheets:
        _build_sheet(
            wb,
            spec,
            with_title_bar=with_title_bar,
            with_frozen_header=with_frozen_header,
            with_glyphs=with_glyphs,
        )

    # openpyxl requires at least one visible sheet to save.
    # If the input was empty, add a minimal placeholder so saving never throws.
    if not wb.sheetnames:
        placeholder = wb.create_sheet("Sheet1")
        placeholder.sheet_state = "visible"

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert markdown to a brand-styled xlsx workbook."
    )
    parser.add_argument("--input", required=True, help="Input markdown file path.")
    parser.add_argument("--output", required=True, help="Output xlsx file path.")
    parser.add_argument(
        "--no-title-bar",
        action="store_true",
        default=False,
        help="Skip the H1 title bar at the top of each sheet.",
    )
    parser.add_argument(
        "--no-frozen-header",
        action="store_true",
        default=False,
        help="Do not freeze the header row.",
    )
    parser.add_argument(
        "--no-glyphs",
        action="store_true",
        default=False,
        help="Do not embed FA glyph icons even if markers are present.",
    )
    args = parser.parse_args()

    md_path = Path(args.input)
    if not md_path.exists():
        raise SystemExit(f"Input file not found: {md_path}")

    md_text = md_path.read_text(encoding="utf-8")

    wb = build_xlsx(
        md_text,
        with_title_bar=not args.no_title_bar,
        with_frozen_header=not args.no_frozen_header,
        with_glyphs=not args.no_glyphs,
    )

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
