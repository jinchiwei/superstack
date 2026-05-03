"""Tests for skills/build-xlsx/build.py.

End-to-end tests: parse fixture markdowns → write xlsx → reload and assert
structure and styling.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

# sys.path is configured by conftest.py in this directory.
_SKILL_DIR = Path(__file__).resolve().parents[1]

import build as bx  # noqa: E402  (import after path manipulation)
from branding_xlsx import (  # type: ignore
    TURQUOISE, DEEPPINK, AMBER, INK, PAPER, WHITE, DEFERRED_GREY
)

# ---------------------------------------------------------------------------
# Fixture paths
# ---------------------------------------------------------------------------
_FIXTURES = Path(__file__).resolve().parent
FIXTURE_SIMPLE   = _FIXTURES / "fixture_simple_table.md"
FIXTURE_MULTI    = _FIXTURES / "fixture_multi_sheet.md"
FIXTURE_MARKERS  = _FIXTURES / "fixture_markers.md"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _bg(cell) -> str:
    """Return the cell's background color as an uppercase hex string (no alpha prefix)."""
    if cell.fill.fgColor.type == "rgb":
        rgb = cell.fill.fgColor.rgb
        # openpyxl stores as AARRGGBB; strip alpha prefix
        return rgb[-6:].upper()
    return ""


def _build(md_text: str, **kwargs):
    """Build a workbook from markdown text and return (Workbook, path)."""
    return bx.build_xlsx(md_text, **kwargs)


def _roundtrip(md_text: str, tmp_path: Path, **kwargs):
    """Build, save, reload, return workbook."""
    import openpyxl
    wb = _build(md_text, **kwargs)
    out = tmp_path / "test.xlsx"
    wb.save(str(out))
    return openpyxl.load_workbook(str(out)), out


# ---------------------------------------------------------------------------
# Basic workbook structure
# ---------------------------------------------------------------------------

class TestWorkbookStructure:
    def test_simple_table_produces_one_sheet(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Results"

    def test_multi_sheet_produces_three_sheets(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_MULTI.read_text(), tmp_path)
        assert len(wb.sheetnames) == 3
        assert "Architecture Sweep" in wb.sheetnames
        assert "Hyperparameter Tuning" in wb.sheetnames
        assert "Future Directions" in wb.sheetnames

    def test_empty_markdown_produces_placeholder_sheet(self, tmp_path):
        """Empty markdown produces a workbook with a single placeholder sheet.

        openpyxl requires at least one visible sheet to save; build_xlsx adds a
        minimal placeholder so callers never encounter an IndexError on save().
        """
        wb, _ = _roundtrip("", tmp_path)
        # Workbook is saveable (no IndexError) and has exactly one sheet
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Sheet1"

    def test_xlsx_file_written(self, tmp_path):
        _, path = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path)
        assert path.exists()
        assert path.stat().st_size > 0


# ---------------------------------------------------------------------------
# Header row styling
# ---------------------------------------------------------------------------

class TestHeaderStyling:
    def test_header_row_ink_background(self, tmp_path):
        """Table header row should have ink-colored background."""
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path)
        ws = wb["Results"]
        # Find the header row (skip title bar in row 1)
        header_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Model":
                    header_row = cell.row
                    break
            if header_row:
                break
        assert header_row is not None, "Header row not found"
        bg = _bg(ws.cell(row=header_row, column=1))
        assert bg == INK.upper(), f"Header bg should be INK ({INK}), got {bg}"

    def test_header_row_white_bold_font(self, tmp_path):
        """Header row font should be white and bold."""
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path)
        ws = wb["Results"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Model":
                    assert cell.font.bold
                    assert cell.font.color.rgb[-6:].upper() == WHITE.upper()
                    return
        pytest.fail("Header cell 'Model' not found")


# ---------------------------------------------------------------------------
# Title bar
# ---------------------------------------------------------------------------

class TestTitleBar:
    def test_title_bar_present_by_default(self, tmp_path):
        """First row of each sheet should be the ink title bar when --no-title-bar not set."""
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path, with_title_bar=True)
        ws = wb["Results"]
        # Row 1 should have ink fill
        bg = _bg(ws.cell(row=1, column=1))
        assert bg == INK.upper(), f"Title bar should have ink bg, got {bg}"

    def test_title_bar_absent_with_no_title_bar(self, tmp_path):
        """With with_title_bar=False, row 1 should be the header row, not ink."""
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path, with_title_bar=False)
        ws = wb["Results"]
        # Row 1 = header row → still ink bg. But table header starts at row 1.
        # The title_bar value should match the header text
        cell = ws.cell(row=1, column=1)
        # Without title bar, first row is the header row (Model, Accuracy...)
        assert cell.value == "Model", f"Expected 'Model' header, got {cell.value!r}"


# ---------------------------------------------------------------------------
# Frozen panes
# ---------------------------------------------------------------------------

class TestFrozenPanes:
    def test_frozen_header_set_by_default(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_SIMPLE.read_text(), tmp_path, with_frozen_header=True)
        ws = wb["Results"]
        assert ws.freeze_panes is not None, "Freeze panes not set"

    def test_no_frozen_header_disables_freeze(self, tmp_path):
        wb, _ = _roundtrip(
            FIXTURE_SIMPLE.read_text(), tmp_path,
            with_frozen_header=False,
        )
        ws = wb["Results"]
        assert ws.freeze_panes is None, "Freeze panes should be None with no_frozen_header"


# ---------------------------------------------------------------------------
# Marker styling
# ---------------------------------------------------------------------------

class TestMarkerStyling:
    def _get_cell_for_value(self, ws, value_fragment: str):
        """Find the first cell whose value contains value_fragment."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and value_fragment in str(cell.value):
                    return cell
        return None

    def test_winner_marker_turquoise_fill(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path)
        ws = wb["Experiment Status"]
        cell = self._get_cell_for_value(ws, "CaFormer + BCE")
        assert cell is not None, "Winner cell not found"
        bg = _bg(cell)
        assert bg == TURQUOISE.upper(), f"Winner should be turquoise, got {bg}"

    def test_deferred_marker_grey_fill(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path)
        ws = wb["Experiment Status"]
        cell = self._get_cell_for_value(ws, "Swin + BCE")
        assert cell is not None, "Deferred cell not found"
        bg = _bg(cell)
        assert bg == DEFERRED_GREY.upper(), f"Deferred should be grey ({DEFERRED_GREY}), got {bg}"

    def test_warning_marker_amber_fill(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path)
        ws = wb["Experiment Status"]
        cell = self._get_cell_for_value(ws, "ViT-B + Focal")
        assert cell is not None, "Warning cell not found"
        bg = _bg(cell)
        assert bg == AMBER.upper(), f"Warning should be amber, got {bg}"

    def test_headline_marker_deeppink_fill(self, tmp_path):
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path)
        ws = wb["Experiment Status"]
        cell = self._get_cell_for_value(ws, "CaFormer + Focal")
        assert cell is not None, "Headline cell not found"
        bg = _bg(cell)
        assert bg == DEEPPINK.upper(), f"Headline should be deeppink, got {bg}"


# ---------------------------------------------------------------------------
# Glyph embedding
# ---------------------------------------------------------------------------

class TestGlyphEmbedding:
    def test_glyphs_embedded_by_default(self, tmp_path):
        """With default settings (with_glyphs=True), images should be present for marker rows."""
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path, with_glyphs=True)
        ws = wb["Experiment Status"]
        # If glyph embedding worked, ws._images should be non-empty
        # (cairosvg must be available for this to work)
        try:
            import cairosvg  # noqa: F401
            has_cairosvg = True
        except ImportError:
            has_cairosvg = False

        if has_cairosvg:
            assert len(ws._images) > 0, "Expected glyph images in worksheet"
        else:
            # Without cairosvg, glyphs gracefully skip — test just checks no crash
            pass

    def test_no_glyphs_flag_suppresses_images(self, tmp_path):
        """With with_glyphs=False, no images should be embedded."""
        wb, _ = _roundtrip(FIXTURE_MARKERS.read_text(), tmp_path, with_glyphs=False)
        ws = wb["Experiment Status"]
        assert len(ws._images) == 0, "No images should be embedded with with_glyphs=False"


# ---------------------------------------------------------------------------
# Prose in non-table sections
# ---------------------------------------------------------------------------

class TestProseRendering:
    def test_prose_in_column_a(self, tmp_path):
        """Non-table H1 sections should render prose text in column A."""
        md = "# Notes\n\nThis is a prose paragraph.\nWith two lines.\n"
        wb, _ = _roundtrip(md, tmp_path, with_title_bar=False)
        ws = wb["Notes"]
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "This is a prose paragraph." in values


# ---------------------------------------------------------------------------
# CLI round-trip
# ---------------------------------------------------------------------------

class TestCLI:
    def test_cli_simple_table(self, tmp_path):
        """Smoke test: CLI invocation produces a valid xlsx."""
        import subprocess
        out_path = tmp_path / "out.xlsx"
        result = subprocess.run(
            [
                sys.executable,
                str(_SKILL_DIR / "build.py"),
                "--input", str(FIXTURE_SIMPLE),
                "--output", str(out_path),
            ],
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0, f"CLI failed: {result.stderr}"
        assert out_path.exists()

    def test_cli_no_glyphs(self, tmp_path):
        """--no-glyphs flag should produce xlsx without images."""
        import openpyxl
        import subprocess
        out_path = tmp_path / "out_noglyphs.xlsx"
        subprocess.run(
            [
                sys.executable,
                str(_SKILL_DIR / "build.py"),
                "--input", str(FIXTURE_MARKERS),
                "--output", str(out_path),
                "--no-glyphs",
            ],
            capture_output=True,
            text=True,
        )
        wb = openpyxl.load_workbook(str(out_path))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert len(ws._images) == 0, f"Sheet {sheet_name} has images but --no-glyphs was set"
